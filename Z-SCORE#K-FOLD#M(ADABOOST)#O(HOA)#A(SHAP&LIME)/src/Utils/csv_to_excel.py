import pandas as pd
from openpyxl import load_workbook

# ✅ Use raw strings to fix escape issues
csv_path = r"D:\ML\Hesam= SCI-6MO-Yahyavi-Deadline=8-7-2025\data\Dataset-6MO-Daily Activity.csv"
excel_path = r"D:\ML\Hesam= SCI-6MO-Yahyavi-Deadline=8-7-2025\data\data.xlsx"

# Step 1: Load CSV file into a DataFrame
df = pd.read_csv(csv_path)

# Step 2: Load existing Excel workbook
book = load_workbook(excel_path)

# Step 3: Append data to Excel
with pd.ExcelWriter(
    excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
) as writer:
    # ✅ No need to set writer.book anymore
    sheet_name = writer.book.sheetnames[0]
    sheet = writer.book[sheet_name]

    # Get the first empty row
    start_row = 0

    # Step 4: Append data without headers
    df.to_excel(
        writer, sheet_name=sheet_name, startrow=start_row, index=False, header=True
    )

print("✅ CSV data has been appended to Excel successfully.")
