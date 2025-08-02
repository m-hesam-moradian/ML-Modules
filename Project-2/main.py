import os
import time
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from src.data_loader import load_data
from src.feature_engineering import average_daily
from src.model.train_model import get_X_y, train_raw_model
from src.Utils.K_Fold import K_Fold
from src.Optimiser.KOA.koa_optimizer import koa_optimizer
from src.Optimiser.COA.coa_optimizer import coa_optimizer
from objective_function import objective_et, objective_lasso

# Timer dictionary
timings = {}

# Load raw data
start = time.time()
DATA_PATH = os.path.join("data", "Data.xlsx")
df = load_data(DATA_PATH)
timings["load_data"] = time.time() - start

# Aggregate daily averages
start = time.time()
df_avg = average_daily(df, samples_per_day=96)
timings["average_daily"] = time.time() - start

# Prepare features and target variable
start = time.time()
X, y = get_X_y(df_avg)
timings["prepare_features"] = time.time() - start

# Apply K-Fold cross-validation
start = time.time()
X_train, X_test, y_train, y_test, etr_scores, lasso_scores, combined_df = K_Fold(
    X, y, n_splits=5
)
timings["k_fold"] = time.time() - start

# Save combined K-Fold data to Excel
start = time.time()
book = load_workbook(DATA_PATH)
if "DATA after K-Fold" in book.sheetnames:
    book.remove(book["DATA after K-Fold"])
    book.save(DATA_PATH)
with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
    combined_df.to_excel(writer, sheet_name="DATA after K-Fold", index=False)
timings["save_to_excel"] = time.time() - start


# Helper function to average metrics
def summarize_metrics(metrics_list):
    return {key: np.mean([m[key] for m in metrics_list]) for key in metrics_list[0]}


avg_etr = summarize_metrics(etr_scores)
avg_lasso = summarize_metrics(lasso_scores)

# Hyperparameter boundaries for ETR and Lasso
lb_et = [10, 1, 2, 1]
ub_et = [200, 50, 10, 10]
dim_et = 4
lb_lasso = [0.1, 100]
ub_lasso = [0.3, 1000]
dim_lasso = 2

n_agents = 3
max_iter = 5

X, y = get_X_y(combined_df)

# Train baseline models
start = time.time()
etr_result = train_raw_model("ETR", X_train, y_train, X_test, y_test)
lr_result = train_raw_model("LR", X_train, y_train, X_test, y_test)
timings["train_baseline_models"] = time.time() - start

# KOA optimization
start = time.time()
best_pos_et_koa, best_fit_et_koa = koa_optimizer(
    objective_et,
    lb_et,
    ub_et,
    dim_et,
    n_agents,
    max_iter,
    X_train,
    y_train,
    X_test,
    y_test,
)
best_pos_lasso_koa, best_fit_lasso_koa = koa_optimizer(
    objective_lasso,
    lb_lasso,
    ub_lasso,
    dim_lasso,
    n_agents,
    max_iter,
    X_train,
    y_train,
    X_test,
    y_test,
)
timings["koa_optimization"] = time.time() - start

# COA optimization
lb_et = [2, 2, 10, 4]
ub_et = [200, 100, 100, 10]
dim_et = 4
lb_lasso = [0.1, 1000]
ub_lasso = [1, 2000]
dim_lasso = 2
n_agents_et = 10
max_iter_et = 20
n_agents_lasso = 100
max_iter_lasso = 20

start = time.time()
best_pos_et_coa, best_fit_et_coa = coa_optimizer(
    objective_et,
    lb_et,
    ub_et,
    dim_et,
    n_agents_et,
    max_iter_et,
    X_train,
    y_train,
    X_test,
    y_test,
)
best_pos_lasso_coa, best_fit_lasso_coa = coa_optimizer(
    objective_lasso,
    lb_lasso,
    ub_lasso,
    dim_lasso,
    n_agents_lasso,
    max_iter_lasso,
    X_train,
    y_train,
    X_test,
    y_test,
)
timings["coa_optimization"] = time.time() - start

# Output timing summary
print("📊 Processing Time Summary (seconds):")
for step, duration in timings.items():
    print(f"  {step}: {duration:.2f}")
