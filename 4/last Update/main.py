import os
import time
import numpy as np
from src.data_loader import load_data
from src.feature_engineering import average_daily
from src.model.train_model import get_X_y, train_raw_model
from src.Utils.K_Fold import K_Fold
from src.Optimiser.KOA.koa_optimizer import koa_optimizer
from src.Optimiser.COA.coa_optimizer import coa_optimizer
from objective_function import objective_et, objective_lasso
from src.Utils.get_optimizer_report import get_optimizer_report
from src.data_loader import load_data
from openpyxl import load_workbook
import pandas as pd

# Timer dictionary
timings = {}


# Load raw data
start = time.time()
DATA_PATH = "./data/Data.xlsx"
df = load_data(DATA_PATH)
timings["load_data"] = time.time() - start

# Aggregate daily averages
start = time.time()
df_avg = average_daily(df, samples_per_day=96)
timings["average_daily"] = time.time() - start

# Prepare features and target variable
start = time.time()
X, y = get_X_y(df_avg)
timings["prepare_features"] = time.time() - start

# Apply K-Fold cross-validation
start = time.time()
X_train, X_test, y_train, y_test, etr_scores, lasso_scores, combined_df = K_Fold(
    X, y, n_splits=5
)
timings["k_fold"] = time.time() - start

# Save combined K-Fold data to Excel
start = time.time()
book = load_workbook(DATA_PATH)
if "DATA after K-Fold" in book.sheetnames:
    book.remove(book["DATA after K-Fold"])
    book.save(DATA_PATH)
with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
    combined_df.to_excel(writer, sheet_name="DATA after K-Fold", index=False)
timings["save_to_excel"] = time.time() - start


# Helper function to average metrics
def summarize_metrics(metrics_list):
    return {key: np.mean([m[key] for m in metrics_list]) for key in metrics_list[0]}


avg_etr = summarize_metrics(etr_scores)
avg_lasso = summarize_metrics(lasso_scores)


# Hyperparameter boundaries for ETR and Lasso
lb_et = [10, 1, 2, 1]
ub_et = [200, 50, 10, 10]
dim_et = 4
lb_lasso = [0.1, 400]
ub_lasso = [0.21, 800]
dim_lasso = 2
n_agents = 5
max_iter = 20


X, y = get_X_y(combined_df)


# Train baseline models
start = time.time()
etr_result = train_raw_model("ETR", X_train, y_train, X_test, y_test)
timings["train_single_ETR"] = time.time() - start
lr_result = train_raw_model("LR", X_train, y_train, X_test, y_test)
timings["train_single_Lasso"] = time.time() - start

# KOA optimization
start = time.time()
best_pos_et_koa, best_fit_et_koa, convergence_et_koa = koa_optimizer(
    objective_et,
    lb_et,
    ub_et,
    dim_et,
    n_agents,
    max_iter,
    X_train,
    y_train,
    X_test,
    y_test,
)
timings["koa_optimization_ETR"] = time.time() - start
best_pos_lasso_koa, best_fit_lasso_koa, convergence_lasso_koa = koa_optimizer(
    objective_lasso,
    lb_lasso,
    ub_lasso,
    dim_lasso,
    n_agents,
    max_iter,
    X_train,
    y_train,
    X_test,
    y_test,
)
timings["koa_optimization_Lasso"] = time.time() - start

# COA optimization


start = time.time()
best_pos_et_coa, best_fit_et_coa, convergence_et_coa = coa_optimizer(
    objective_et,
    lb_et,
    ub_et,
    dim_et,
    n_agents,
    max_iter,
    X_train,
    y_train,
    X_test,
    y_test,
)
timings["coa_optimization_ETR"] = time.time() - start
best_pos_lasso_coa, best_fit_lasso_coa, convergence_lasso_coa = coa_optimizer(
    objective_lasso,
    lb_lasso,
    ub_lasso,
    dim_lasso,
    n_agents,
    max_iter,
    X_train,
    y_train,
    X_test,
    y_test,
)
timings["coa_optimization_Lasso"] = time.time() - start


optimizer_results = {}
# 🧪 Run and store optimizer reports
optimizer_results["etr_koa"] = get_optimizer_report(
    "ETR_KOA",
    "ETR",
    best_pos_et_koa,
    best_fit_et_koa,
    X_train,
    y_train,
    X_test,
    y_test,
    convergence_et_koa,
)

optimizer_results["lasso_koa"] = get_optimizer_report(
    "Lasso_KOA",
    "LR",
    best_pos_lasso_koa,
    best_fit_lasso_koa,
    X_train,
    y_train,
    X_test,
    y_test,
    convergence_lasso_koa,
)


optimizer_results["etr_coa"] = get_optimizer_report(
    "ETR_COA",
    "ETR",
    best_pos_et_coa,
    best_fit_et_coa,
    X_train,
    y_train,
    X_test,
    y_test,
    convergence_et_coa,
)

optimizer_results["lasso_coa"] = get_optimizer_report(
    "Lasso_COA",
    "LR",
    best_pos_lasso_coa,
    best_fit_lasso_coa,
    X_train,
    y_train,
    X_test,
    y_test,
    convergence_lasso_coa,
)
# Output timing summary
print("📊 Processing Time Summary (seconds):")
for step, duration in timings.items():
    print(f"  {step}: {duration:.2f}")

# Output average metrics
