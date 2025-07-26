import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

def write_metrics_table(ws, metrics_dict, start_row, start_col, fill_color, label):
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    n_cols = len(metrics_dict)
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + n_cols - 1)
    label_cell = ws.cell(row=start_row, column=start_col, value=label)
    label_cell.font = Font(bold=True, color="FFFFFF")
    label_cell.alignment = center_align
    label_cell.fill = fill

    # Headers
    header_row = start_row + 1
    for i, header in enumerate(metrics_dict.keys()):
        cell = ws.cell(row=header_row, column=start_col + i, value=header.capitalize())
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = fill

    # Values
    value_row = header_row + 1
    for i, value in enumerate(metrics_dict.values()):
        cell = ws.cell(row=value_row, column=start_col + i, value=round(value, 4))
        cell.alignment = center_align


def save_predictions_and_metrics_to_excel(results_dict, model_name, excel_path, sheet_name=None):
    sheet_name = sheet_name or model_name

    result_all = results_dict[model_name]["all"]
    result_train = results_dict[model_name]["train"]
    result_test = results_dict[model_name]["test"]

    n_train = len(result_train["true_values"])
    all_true = result_all["true_values"]
    all_preds = result_all["predictions"]
    data_labels = ["train"] * n_train + ["test"] * (len(all_true) - n_train)

    # 1. Save predictions
    df_combined = pd.DataFrame({
        "ID": range(1, len(all_true) + 1),
        "True": all_true,
        "Predicted": all_preds,
        "Set": data_labels
    })

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_combined.drop(columns=["Set"]).to_excel(writer, sheet_name=sheet_name, index=False)

    # 2. Style in openpyxl
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # Styling
    train_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # light green
    test_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # light red
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    header = [cell.value for cell in ws[1]]
    predicted_col_index = header.index("Predicted") + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        for col_idx, cell in enumerate(row, start=1):
            cell.alignment = center_align
            if col_idx == predicted_col_index:
                fill = train_fill if data_labels[row_idx] == "train" else test_fill
                cell.fill = fill

    for cell in ws[1]:
        cell.alignment = center_align
        cell.font = bold_font

    # 3. Add metric tables
    start_col = ws.max_column + 2
    start_row = 1

    write_metrics_table(ws, 
                        {k: v for k, v in result_all.items() if k not in ["predictions", "true_values"]}, 
                        start_row, start_col, "ADD8E6", "All Metrics")

    write_metrics_table(ws, 
                        {k: v for k, v in result_train.items() if k not in ["predictions", "true_values"]}, 
                        start_row + 5, start_col, "D0F0C0", "Train Metrics")

    write_metrics_table(ws, 
                        {k: v for k, v in result_test.items() if k not in ["predictions", "true_values"]}, 
                        start_row + 10, start_col, "FADADD", "Test Metrics")

    wb.save(excel_path)
