import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from src.Utils.K_Fold import K_Fold
from src.data_loader import load_data
from src.analysis.SHAP import shap_analysis
from sklearn.ensemble import AdaBoostRegressor
from src.feature_engineering import average_daily
from objective_function import objective_adaboost
from src.Utils.Z_Score import remove_outliers_zscore
from src.model.train_model import get_X_y, train_model
from src.analysis.LIME import lime_sensitivity_analysis
from src.Optimiser.HOA.hoa_optimizer import hoa_optimizer


# Load raw data
DATA_PATH = "D:/ML/Hesam= SCI-6MO-Yahyavi-Deadline=8-7-2025/data/data.xlsx"

df = load_data(DATA_PATH)


# Remove outliers using the function
cleaned_df, z_scores = remove_outliers_zscore(df, threshold=3)

# Display original and cleaned DataFrame
# book = load_workbook(DATA_PATH)
# if "Data after Z-score" in book.sheetnames:
#     book.remove(book["Data after Z-score"])
#     book.save(DATA_PATH)
# with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
#     cleaned_df.to_excel(writer, sheet_name="Data after Z-score", index=False)

X, y = get_X_y(cleaned_df, target_col="Burned Calories ")


# Apply K-Fold cross-validation
(
    X_train_best,
    X_test_best,
    y_train_best,
    y_test_best,
    K_Fold_Cross_Validation_Scores,
    combined_df,
) = K_Fold(X, y, n_splits=5)

K_Fold_Cross_Validation_Scores = pd.DataFrame(K_Fold_Cross_Validation_Scores)

# Save combined K-Fold data to Excel
# book = load_workbook(DATA_PATH)
# if "DATA after K-Fold" in book.sheetnames:
#     book.remove(book["DATA after K-Fold"])
#     book.save(DATA_PATH)
# with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
#     combined_df.to_excel(writer, sheet_name="DATA after K-Fold", index=False)


# # Helper function to average metrics
# def summarize_metrics(metrics_list):
#     return {key: np.mean([m[key] for m in metrics_list]) for key in metrics_list[0]}


# avg_metrics_k_fold = summarize_metrics(K_Fold_Cross_Validation_Scores)
# print("Average Metrics:")
# for key, value in avg_metrics_k_fold.items():
#     print(f"{key}: {value:.4f}")


# singleModel_result = train_model(X_train_best, y_train_best, X_test_best, y_test_best)


# best_pos, best_RMSE, convergence = hoa_optimizer(
#     objective_adaboost,  # our AdaBoost objective
#     [500, 0.01],  # lower bounds: n_estimators, learning_rate
#     [1000, 1.0],  # upper bounds
#     2,  # dim
#     5,  # n_agents
#     5,  # max_iter
#     X_train_best,
#     y_train_best,
#     X_test_best,
#     y_test_best,
# )

# HOA_model_result = train_model(
#     X_train_best, y_train_best, X_test_best, y_test_best, best_pos
# )

model = AdaBoostRegressor()
model.fit(X_train_best, y_train_best)

# # SHAP on the HOA model
# sensitivity_df_shap, shap_values = shap_analysis(
#     model=model,
#     X_train=X_train_best,
#     y_train=y_train_best,
#     X_test=X_test_best,
#     y_test=y_test_best,
#     save_path=DATA_PATH,  # save to same Excel file
#     sheet_name="SHAP_Sensitivity",
# )


sensitivity_LIME = pd.DataFrame([
    lime_sensitivity_analysis(
        model=model,
        X_train=X_train_best,
        y_train=y_train_best,
        X_test=X_test_best,
        y_test=y_test_best,
        sample_index=5,
        epsilon=0.05,
    )
])


# print("\nBest AdaBoost Params:", best_pos)
# print("Best RMSE:", best_RMSE)
